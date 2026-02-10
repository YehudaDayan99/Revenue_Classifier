#!/usr/bin/env python3
"""Create Excel file from company list."""
import pandas as pd
from pathlib import Path

companies = [
    ["Apple Inc.", "AAPL", "Semiconductors"],
    ["Microsoft Corporation", "MSFT", "Software – Infrastructure"],
    ["NVIDIA Corporation", "NVDA", "Semiconductors"],
    ["Amazon.com, Inc.", "AMZN", "Internet & Direct Marketing Retail"],
    ["Alphabet Inc.", "GOOGL", "Internet Content & Information Providers"],
    ["Saudi Arabian Oil Company", "2222.SR", "Oil & Gas Exploration & Production"],
    ["Meta Platforms, Inc.", "META", "Interactive Media & Services"],
    ["Berkshire Hathaway Inc.", "BRK.B", "Multi‑Sector Holdings"],
    ["Tesla, Inc.", "TSLA", "Auto Manufacturers"],
    ["Broadcom Inc.", "AVGO", "Semiconductors"],
    ["Eli Lilly and Company", "LLY", "Pharmaceuticals"],
    ["Taiwan Semiconductor Manufacturing Company Limited", "2330.TW", "Semiconductors"],
    ["Walmart Inc.", "WMT", "Hypermarkets & Super Centers"],
    ["JPMorgan Chase & Co.", "JPM", "Diversified Banks"],
    ["Visa Inc.", "V", "Transaction & Payment Processing Services"],
    ["Tencent Holdings Limited", "0700.HK", "Internet Software & Services"],
    ["Exxon Mobil Corporation", "XOM", "Oil & Gas Exploration & Production"],
    ["Mastercard Incorporated", "MA", "Transaction & Payment Processing Services"],
    ["UnitedHealth Group Incorporated", "UNH", "Managed Health Care"],
    ["Costco Wholesale Corporation", "COST", "Membership Retailers"],
    ["Johnson & Johnson", "JNJ", "Pharmaceuticals"],
    ["The Procter & Gamble Company", "PG", "Household Products"],
    ["Netflix, Inc.", "NFLX", "Entertainment"],
    ["Oracle Corporation", "ORCL", "Software – Application"],
    ["AbbVie Inc.", "ABBV", "Pharmaceuticals"],
    ["The Home Depot, Inc.", "HD", "Home Improvement Retail"],
    ["Industrial and Commercial Bank of China Limited", "1398.HK", "Diversified Banks"],
    ["Bank of America Corporation", "BAC", "Diversified Banks"],
    ["LVMH Moët Hennessy - Louis Vuitton", "MC.PA", "Luxury Goods"],
    ["The Coca‑Cola Company", "KO", "Beverages – Soft Drinks"],
    ["SAP SE", "SAP.DE", "Software – Application"],
    ["T‑Mobile US, Inc.", "TMUS", "Wireless Telecommunication Services"],
    ["Novo Nordisk A/S", "NOVO B", "Pharmaceuticals"],
    ["Alibaba Group Holding Limited", "09988.HK", "Internet Software & Services"],
    ["Chevron Corporation", "CVX", "Oil & Gas Exploration & Production"],
    ["Hermès International", "RMS.PA", "Luxury Goods"],
    ["Kweichow Moutai Co., Ltd.", "600519.SS", "Beverages – Distilled & Wine"],
    ["Roche Holding AG", "Roche Holding AG", "Pharmaceuticals"],
    ["Nestlé S.A.", "NESN.SW", "Food Products"],
    ["Samsung Electronics Co., Ltd.", "005930.KS", "Semiconductors"],
    ["Salesforce, Inc.", "CRM", "Software – Application"],
    ["ASML Holding N.V.", "ASML", "Specialized Semiconductor Equipment"],
    ["Philip Morris International Inc.", "PM", "Tobacco"],
    ["Agricultural Bank of China Limited", "1288.HK", "Diversified Banks"],
    ["Cisco Systems, Inc.", "CSCO", "Communications Equipment"],
    ["International Holding Company", "IHC.AD", "Industrial Conglomerates"],
    ["China Mobile Limited", "0941.HK", "Wireless Telecommunication Services"],
    ["Wells Fargo & Company", "WFC", "Diversified Banks"],
    ["International Business Machines Corporation", "IBM", "IT Consulting & Other Services"],
    ["Abbott Laboratories", "ABT", "Health Care Equipment"],
    ["Toyota Motor Corporation", "TM", "Auto Manufacturers"],
    ["Merck & Co., Inc.", "MRK", "Pharmaceuticals"],
    ["AstraZeneca PLC", "AZN", "Pharmaceuticals"],
    ["China Construction Bank Corporation", "0939.HK", "Diversified Banks"],
    ["McDonald's Corporation", "MCD", "Restaurants"],
    ["Shell plc", "SHEL", "Oil & Gas Exploration & Production"],
    ["Linde plc", "LIN", "Industrial Gases"],
    ["General Electric Company", "GE", "Industrial Conglomerates"],
    ["Bank of China Limited", "3988.HK", "Diversified Banks"],
    ["Novartis AG", "NVS", "Pharmaceuticals"],
    ["PepsiCo, Inc.", "PEP", "Beverages – Soft Drinks"],
    ["AT&T Inc.", "T", "Integrated Telecommunication Services"],
    ["PetroChina Company Limited", "0857.HK", "Oil & Gas Exploration & Production"],
    ["HSBC Holdings plc", "HSBA.L", "Diversified Banks"],
    ["L'Oréal S.A.", "OR.PA", "Personal Products"],
    ["Palantir Technologies Inc.", "PLTR", "Application Software"],
    ["Accenture plc", "ACN", "IT Consulting & Other Services"],
    ["Verizon Communications Inc.", "VZ", "Integrated Telecommunication Services"],
    ["Reliance Industries Limited", "RELIANCE.NS", "Integrated Oil & Gas"],
    ["American Express Company", "AXP", "Consumer Finance"],
    ["Morgan Stanley", "MS", "Diversified Banks"],
    ["Thermo Fisher Scientific Inc.", "TMO", "Health Care Equipment"],
    ["Deutsche Telekom AG", "DTE.DE", "Integrated Telecommunication Services"],
    ["Siemens Aktiengesellschaft", "SIE.DE", "Industrial Machinery"],
    ["The Walt Disney Company", "DIS", "Entertainment"],
    ["Intuitive Surgical, Inc.", "ISRG", "Health Care Equipment"],
    ["The Goldman Sachs Group, Inc.", "GS", "Diversified Banks"],
    ["RTX Corporation", "RTX", "Aerospace & Defense"],
    ["Intuit Inc.", "INTU", "Software – Application"],
    ["QUALCOMM Incorporated", "QCOM", "Semiconductors"],
    ["Amgen Inc.", "AMGN", "Biotechnology"],
    ["Advanced Micro Devices, Inc.", "AMD", "Semiconductors"],
    ["The Progressive Corporation", "PGR", "Property & Casualty Insurance"],
    ["PDD Holdings Inc.", "PDD", "Internet Software & Services"],
    ["ServiceNow, Inc.", "NOW", "Software – Application"],
    ["HDFC Bank Limited", "HDFCBANK.NS", "Diversified Banks"],
    ["Texas Instruments Incorporated", "TXN", "Semiconductors"],
    ["Adobe Inc.", "ADBE", "Software – Application"],
    ["Royal Bank of Canada", "RY", "Diversified Banks"],
    ["Xiaomi Corporation", "1810.HK", "Technology Hardware, Storage & Peripherals"],
    ["Caterpillar Inc.", "CAT", "Construction Machinery & Heavy Trucks"],
    ["Commonwealth Bank of Australia", "CBA.AX", "Diversified Banks"],
    ["S&P Global Inc.", "SPGI", "Financial Exchanges & Data"],
    ["BYD Company Limited", "BYDDY", "Auto Manufacturers"],
    ["Mitsubishi UFJ Financial Group, Inc.", "MUFG", "Diversified Banks"],
    ["Industria de Diseño Textil, S.A.", "ITX", "Apparel Retail"],
    ["Contemporary Amperex Technology Co., Limited", "CATL", "Electrical Components & Equipment"],
    ["Tata Consultancy Services Limited", "TCS.NS", "IT Consulting & Other Services"],
    ["Uber Technologies, Inc.", "UBER", "Internet Software & Services"],
    ["Sony Group Corporation", "SONY", "Consumer Electronics"],
]

df = pd.DataFrame(companies, columns=["Name", "Ticker", "GICS_Level_4"])
output_dir = Path("data/lists")
output_dir.mkdir(parents=True, exist_ok=True)
output_file = output_dir / "company_list.xlsx"

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Companies', index=False)
    worksheet = writer.sheets['Companies']
    
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    
    for idx, col in enumerate(df.columns, 1):
        max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2
        worksheet.column_dimensions[get_column_letter(idx)].width = max_length
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

print(f"✓ Created: {output_file}")
print(f"  {len(df)} companies")
